from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from core.roles_decorators import permiso_required
from django.db.models import Q, Sum, Count, Avg
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from datetime import timedelta, date, datetime
from calendar import monthrange
import json
import logging
import requests
from .models import Pago, PlanPago, TransaccionPago
from .forms import PagoForm, PlanPagoForm
from .payment_gateway import PaymentGateway
from clientes.models import Cliente
from instalaciones.models import Instalacion

logger = logging.getLogger(__name__)


@login_required
@permiso_required('ver_pagos')
def pago_list(request):
    """Lista todos los pagos con búsqueda y filtros."""
    # Actualizar automáticamente pagos vencidos
    Pago.actualizar_pagos_vencidos()
    
    pagos = Pago.objects.select_related('cliente', 'instalacion').all()
    
    # Búsqueda
    query = request.GET.get('q', '')
    if query:
        pagos = pagos.filter(
            Q(cliente__nombre__icontains=query) |
            Q(cliente__apellido1__icontains=query) |
            Q(cliente__apellido2__icontains=query) |
            Q(cliente__telefono__icontains=query) |
            Q(concepto__icontains=query) |
            Q(referencia_pago__icontains=query)
        )
    
    # Filtro por estado
    estado_filter = request.GET.get('estado', '')
    if estado_filter:
        pagos = pagos.filter(estado=estado_filter)
    
    # Filtro por método de pago
    metodo_filter = request.GET.get('metodo', '')
    if metodo_filter:
        pagos = pagos.filter(metodo_pago=metodo_filter)
    
    # Filtro por período
    periodo_anio = request.GET.get('anio', '')
    periodo_mes = request.GET.get('mes', '')
    if periodo_anio:
        pagos = pagos.filter(periodo_anio=periodo_anio)
    if periodo_mes:
        pagos = pagos.filter(periodo_mes=periodo_mes)
    
    # Ordenamiento
    orden = request.GET.get('orden', '-fecha_vencimiento')
    pagos = pagos.order_by(orden)
    
    # Calcular estadísticas
    total_pagos = pagos.count()
    total_monto = pagos.aggregate(Sum('monto'))['monto__sum'] or 0
    pagos_pendientes = pagos.filter(estado='pendiente').count()
    pagos_vencidos = pagos.filter(estado='vencido').count()
    pagos_pagados = pagos.filter(estado='pagado').count()
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(pagos, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'query': query,
        'estado_filter': estado_filter,
        'metodo_filter': metodo_filter,
        'periodo_anio': periodo_anio,
        'periodo_mes': periodo_mes,
        'orden': orden,
        'estados': Pago.ESTADO_CHOICES,
        'metodos': Pago.METODO_PAGO_CHOICES,
        'total_pagos': total_pagos,
        'total_monto': total_monto,
        'pagos_pendientes': pagos_pendientes,
        'pagos_vencidos': pagos_vencidos,
        'pagos_pagados': pagos_pagados,
    }
    
    return render(request, 'pagos/pago_list.html', context)


@login_required
@permiso_required('ver_pagos')
def pago_detail(request, pk):
    """Muestra los detalles de un pago."""
    pago = get_object_or_404(Pago.objects.select_related('cliente', 'instalacion'), pk=pk)
    
    # Obtener notificaciones relacionadas si existen
    notificaciones = pago.notificaciones.all() if hasattr(pago, 'notificaciones') else []
    
    # Obtener transacciones de pasarela
    transacciones = pago.transacciones.all().order_by('-fecha_creacion')
    
    # Verificar si hay alguna pasarela configurada
    tiene_pasarela = (
        bool(getattr(settings, 'STRIPE_SECRET_KEY', None)) or
        bool(getattr(settings, 'MERCADOPAGO_ACCESS_TOKEN', None)) or
        (bool(getattr(settings, 'PAYPAL_CLIENT_ID', None)) and bool(getattr(settings, 'PAYPAL_SECRET', None)))
    )
    
    context = {
        'pago': pago,
        'notificaciones': notificaciones,
        'transacciones': transacciones,
        'tiene_pasarela': tiene_pasarela,
    }
    
    return render(request, 'pagos/pago_detail.html', context)


@login_required
@permiso_required('crear_pagos')
def pago_create(request, cliente_id=None):
    """Crea un nuevo pago."""
    cliente_pre_seleccionado = None
    if cliente_id:
        try:
            cliente_pre_seleccionado = Cliente.objects.get(pk=cliente_id)
        except Cliente.DoesNotExist:
            messages.error(request, 'Cliente no encontrado.')
            return redirect('pagos:pago_list')
    
    if request.method == 'POST':
        form = PagoForm(request.POST, cliente_id=cliente_id)
        if form.is_valid():
            pago = form.save()
            messages.success(request, f'Pago de ${pago.monto} creado exitosamente.')
            return redirect('pagos:pago_detail', pk=pago.pk)
    else:
        form = PagoForm(cliente_id=cliente_id)
        if cliente_pre_seleccionado:
            form.fields['cliente'].initial = cliente_pre_seleccionado
            form.fields['instalacion'].queryset = cliente_pre_seleccionado.instalaciones.all()
    
    context = {
        'form': form,
        'title': 'Nuevo Pago',
        'cliente_pre_seleccionado': cliente_pre_seleccionado,
    }
    
    return render(request, 'pagos/pago_form.html', context)


@login_required
@permiso_required('editar_pagos')
def pago_update(request, pk):
    """Actualiza un pago existente."""
    pago = get_object_or_404(Pago, pk=pk)
    
    if request.method == 'POST':
        form = PagoForm(request.POST, instance=pago)
        if form.is_valid():
            pago = form.save()
            messages.success(request, f'Pago actualizado exitosamente.')
            return redirect('pagos:pago_detail', pk=pago.pk)
    else:
        form = PagoForm(instance=pago)
        # Cargar instalaciones del cliente si existe
        if pago.cliente:
            form.fields['instalacion'].queryset = pago.cliente.instalaciones.all()
    
    context = {
        'form': form,
        'pago': pago,
        'title': 'Editar Pago',
        'cliente_pre_seleccionado': pago.cliente,  # Para mostrar en el buscador
    }
    
    return render(request, 'pagos/pago_form.html', context)


@login_required
@permiso_required('eliminar_pagos')
def pago_delete(request, pk):
    """Elimina un pago."""
    pago = get_object_or_404(Pago, pk=pk)
    
    if request.method == 'POST':
        pago_str = f"${pago.monto} - {pago.cliente.nombre_completo}"
        pago.delete()
        messages.success(request, f'Pago "{pago_str}" eliminado exitosamente.')
        return redirect('pagos:pago_list')
    
    context = {
        'pago': pago,
    }
    
    return render(request, 'pagos/pago_confirm_delete.html', context)


@login_required
@permiso_required('marcar_pagos_pagados')
def pago_marcar_pagado(request, pk):
    """Marca un pago como pagado."""
    pago = get_object_or_404(Pago, pk=pk)
    
    if request.method == 'POST':
        metodo_pago = request.POST.get('metodo_pago', '')
        referencia_pago = request.POST.get('referencia_pago', '')
        
        pago.marcar_como_pagado(
            metodo_pago=metodo_pago if metodo_pago else None,
            referencia=referencia_pago if referencia_pago else None
        )
        messages.success(request, f'Pago marcado como pagado exitosamente.')
        return redirect('pagos:pago_detail', pk=pago.pk)
    
    context = {
        'pago': pago,
        'metodos': Pago.METODO_PAGO_CHOICES,
    }
    
    return render(request, 'pagos/pago_marcar_pagado.html', context)


# ============================================
# API para búsqueda de clientes
# ============================================

@login_required
def buscar_clientes(request):
    """API para buscar clientes por nombre, apellido, teléfono o email."""
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'clientes': []})
    
    clientes = Cliente.objects.filter(
        Q(nombre__icontains=query) |
        Q(apellido1__icontains=query) |
        Q(apellido2__icontains=query) |
        Q(telefono__icontains=query) |
        Q(email__icontains=query)
    ).order_by('nombre', 'apellido1')[:15]
    
    data = {
        'clientes': [
            {
                'id': c.id,
                'nombre_completo': c.nombre_completo,
                'telefono': c.telefono,
                'email': c.email or '',
                'estado': c.get_estado_cliente_display(),
                'ciudad': c.ciudad or '',
            }
            for c in clientes
        ]
    }
    
    return JsonResponse(data)


@login_required
def obtener_instalaciones_cliente(request, cliente_id):
    """API para obtener las instalaciones de un cliente con información de PlanPago."""
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
        instalaciones = cliente.instalaciones.all()
        
        instalaciones_data = []
        for inst in instalaciones:
            # Obtener PlanPago si existe
            plan_pago = None
            try:
                plan_pago = inst.plan_pago
            except:
                pass
            
            instalacion_data = {
                'id': inst.id,
                'plan_nombre': inst.plan_nombre or f'Instalación #{inst.id}',
                'estado': inst.get_estado_display(),
                'precio_mensual': float(inst.precio_mensual) if inst.precio_mensual else None,
                'direccion': inst.direccion_instalacion or '',
            }
            
            # Agregar información del PlanPago si existe
            if plan_pago and plan_pago.activo:
                instalacion_data['plan_pago'] = {
                    'monto_mensual': float(plan_pago.monto_mensual),
                    'dia_vencimiento': plan_pago.dia_vencimiento,
                    'activo': plan_pago.activo,
                }
            
            instalaciones_data.append(instalacion_data)
        
        data = {
            'cliente': {
                'id': cliente.id,
                'nombre_completo': cliente.nombre_completo,
            },
            'instalaciones': instalaciones_data
        }
        
        return JsonResponse(data)
    except Cliente.DoesNotExist:
        return JsonResponse({'error': 'Cliente no encontrado'}, status=404)


# ============================================
# Exportación de Pagos
# ============================================

@login_required
def pago_exportar_excel(request):
    """Exporta los pagos filtrados a Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'La librería openpyxl no está instalada. Ejecuta: pip install openpyxl')
        return redirect('pagos:pago_list')
    
    # Obtener los mismos filtros que en pago_list
    pagos = Pago.objects.select_related('cliente', 'instalacion').all()
    
    query = request.GET.get('q', '')
    if query:
        pagos = pagos.filter(
            Q(cliente__nombre__icontains=query) |
            Q(cliente__apellido1__icontains=query) |
            Q(cliente__apellido2__icontains=query) |
            Q(cliente__telefono__icontains=query) |
            Q(concepto__icontains=query) |
            Q(referencia_pago__icontains=query)
        )
    
    estado_filter = request.GET.get('estado', '')
    if estado_filter:
        pagos = pagos.filter(estado=estado_filter)
    
    metodo_filter = request.GET.get('metodo', '')
    if metodo_filter:
        pagos = pagos.filter(metodo_pago=metodo_filter)
    
    periodo_anio = request.GET.get('anio', '')
    periodo_mes = request.GET.get('mes', '')
    if periodo_anio:
        pagos = pagos.filter(periodo_anio=periodo_anio)
    if periodo_mes:
        pagos = pagos.filter(periodo_mes=periodo_mes)
    
    orden = request.GET.get('orden', '-fecha_vencimiento')
    pagos = pagos.order_by(orden)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Pagos"
    
    # Estilos
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Encabezados
    headers = ['Cliente', 'Instalación', 'Concepto', 'Monto', 'Período', 
               'Fecha Vencimiento', 'Fecha Pago', 'Estado', 'Método Pago', 
               'Referencia', 'Notas']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    for row, pago in enumerate(pagos, 2):
        ws.cell(row=row, column=1, value=pago.cliente.nombre_completo)
        ws.cell(row=row, column=2, value=pago.instalacion.plan_nombre if pago.instalacion else '')
        ws.cell(row=row, column=3, value=pago.concepto)
        ws.cell(row=row, column=4, value=float(pago.monto))
        ws.cell(row=row, column=5, value=f"{pago.get_periodo_mes_display()} {pago.periodo_anio}")
        ws.cell(row=row, column=6, value=pago.fecha_vencimiento.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=7, value=pago.fecha_pago.strftime('%d/%m/%Y %H:%M') if pago.fecha_pago else '')
        ws.cell(row=row, column=8, value=pago.get_estado_display())
        ws.cell(row=row, column=9, value=pago.get_metodo_pago_display() if pago.metodo_pago else '')
        ws.cell(row=row, column=10, value=pago.referencia_pago or '')
        ws.cell(row=row, column=11, value=pago.notas or '')
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'pagos_export_{date.today().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
@permiso_required('ver_pagos')
def pago_exportar_pdf(request):
    """Exporta los pagos filtrados a PDF."""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        messages.error(request, 'La librería reportlab no está instalada. Ejecuta: pip install reportlab')
        return redirect('pagos:pago_list')
    
    # Obtener los mismos filtros que en pago_list
    pagos = Pago.objects.select_related('cliente', 'instalacion').all()
    
    query = request.GET.get('q', '')
    if query:
        pagos = pagos.filter(
            Q(cliente__nombre__icontains=query) |
            Q(cliente__apellido1__icontains=query) |
            Q(cliente__apellido2__icontains=query) |
            Q(cliente__telefono__icontains=query) |
            Q(concepto__icontains=query) |
            Q(referencia_pago__icontains=query)
        )
    
    estado_filter = request.GET.get('estado', '')
    if estado_filter:
        pagos = pagos.filter(estado=estado_filter)
    
    metodo_filter = request.GET.get('metodo', '')
    if metodo_filter:
        pagos = pagos.filter(metodo_pago=metodo_filter)
    
    periodo_anio = request.GET.get('anio', '')
    periodo_mes = request.GET.get('mes', '')
    if periodo_anio:
        pagos = pagos.filter(periodo_anio=periodo_anio)
    if periodo_mes:
        pagos = pagos.filter(periodo_mes=periodo_mes)
    
    orden = request.GET.get('orden', '-fecha_vencimiento')
    pagos = pagos.order_by(orden)
    
    # Crear respuesta PDF
    response = HttpResponse(content_type='application/pdf')
    filename = f'pagos_export_{date.today().strftime("%Y%m%d")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Crear documento
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=30,
    )
    
    # Título
    elements.append(Paragraph('Reporte de Pagos', title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Información del reporte
    info_text = f"<b>Fecha de generación:</b> {date.today().strftime('%d/%m/%Y')}<br/>"
    info_text += f"<b>Total de pagos:</b> {pagos.count()}<br/>"
    if query:
        info_text += f"<b>Búsqueda:</b> {query}<br/>"
    if estado_filter:
        info_text += f"<b>Estado:</b> {dict(Pago.ESTADO_CHOICES).get(estado_filter, estado_filter)}<br/>"
    
    elements.append(Paragraph(info_text, styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Preparar datos de la tabla
    data = [['Cliente', 'Concepto', 'Monto', 'Período', 'Vencimiento', 'Estado']]
    
    for pago in pagos[:100]:  # Limitar a 100 para no sobrecargar el PDF
        data.append([
            pago.cliente.nombre_completo[:30],  # Truncar nombres largos
            pago.concepto[:25],
            f"${pago.monto:.2f}",
            f"{pago.get_periodo_mes_display()} {pago.periodo_anio}",
            pago.fecha_vencimiento.strftime('%d/%m/%Y'),
            pago.get_estado_display()
        ])
    
    # Crear tabla
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    return response


# ============================================
# Vista de Calendario
# ============================================

@login_required
@permiso_required('ver_pagos')
def pago_calendario(request):
    """Vista de calendario mensual de pagos."""
    # Obtener mes y año de la URL o usar el actual
    hoy = date.today()
    try:
        anio = int(request.GET.get('anio', hoy.year))
        mes = int(request.GET.get('mes', hoy.month))
    except (ValueError, TypeError):
        anio = hoy.year
        mes = hoy.month
    
    # Validar mes y año
    if mes < 1 or mes > 12:
        mes = hoy.month
    if anio < 2000 or anio > 2100:
        anio = hoy.year
    
    # Obtener pagos del mes
    pagos = Pago.objects.filter(
        fecha_vencimiento__year=anio,
        fecha_vencimiento__month=mes
    ).select_related('cliente', 'instalacion').order_by('fecha_vencimiento')
    
    # Organizar pagos por día
    pagos_por_dia = {}
    for pago in pagos:
        dia = pago.fecha_vencimiento.day
        if dia not in pagos_por_dia:
            pagos_por_dia[dia] = []
        pagos_por_dia[dia].append(pago)
    
    # Calcular días del mes y primer día de la semana
    dias_en_mes = monthrange(anio, mes)[1]
    primer_dia = date(anio, mes, 1)
    primer_dia_semana = primer_dia.weekday()  # 0 = Lunes, 6 = Domingo
    
    # Meses en español
    meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    
    # Calcular mes anterior y siguiente
    if mes == 1:
        mes_anterior = 12
        anio_anterior = anio - 1
    else:
        mes_anterior = mes - 1
        anio_anterior = anio
    
    if mes == 12:
        mes_siguiente = 1
        anio_siguiente = anio + 1
    else:
        mes_siguiente = mes + 1
        anio_siguiente = anio
    
    # Estadísticas del mes
    total_pagos = pagos.count()
    total_monto = pagos.aggregate(Sum('monto'))['monto__sum'] or 0
    pagos_pendientes = pagos.filter(estado='pendiente').count()
    pagos_vencidos = pagos.filter(estado='vencido').count()
    pagos_pagados = pagos.filter(estado='pagado').count()
    
    context = {
        'anio': anio,
        'mes': mes,
        'mes_nombre': meses[mes],
        'dias_en_mes': dias_en_mes,
        'primer_dia_semana': primer_dia_semana,
        'pagos_por_dia': pagos_por_dia,
        'mes_anterior': mes_anterior,
        'anio_anterior': anio_anterior,
        'mes_siguiente': mes_siguiente,
        'anio_siguiente': anio_siguiente,
        'total_pagos': total_pagos,
        'total_monto': total_monto,
        'pagos_pendientes': pagos_pendientes,
        'pagos_vencidos': pagos_vencidos,
        'pagos_pagados': pagos_pagados,
        'hoy': hoy,
    }
    
    return render(request, 'pagos/pago_calendario.html', context)


# ============================================
# Reportes Financieros
# ============================================

@login_required
@permiso_required('ver_reportes_pagos')
def pago_reportes(request):
    """Vista de reportes financieros."""
    # Obtener año de la URL o usar el actual
    hoy = date.today()
    try:
        anio = int(request.GET.get('anio', hoy.year))
    except (ValueError, TypeError):
        anio = hoy.year
    
    # Generar lista de años disponibles (2020-2030)
    años_disponibles = list(range(2020, 2031))
    
    # Validar año
    if anio < 2000 or anio > 2100:
        anio = hoy.year
    
    # Obtener todos los pagos del año
    pagos = Pago.objects.filter(periodo_anio=anio).select_related('cliente', 'instalacion')
    
    # Estadísticas generales
    total_pagos = pagos.count()
    total_monto = pagos.aggregate(Sum('monto'))['monto__sum'] or 0
    monto_pagado = pagos.filter(estado='pagado').aggregate(Sum('monto'))['monto__sum'] or 0
    monto_pendiente = pagos.filter(estado__in=['pendiente', 'vencido']).aggregate(Sum('monto'))['monto__sum'] or 0
    
    # Ingresos por mes
    ingresos_por_mes = []
    for mes in range(1, 13):
        pagos_mes = pagos.filter(periodo_mes=mes, estado='pagado')
        monto_mes = pagos_mes.aggregate(Sum('monto'))['monto__sum'] or 0
        cantidad_mes = pagos_mes.count()
        ingresos_por_mes.append({
            'mes': mes,
            'mes_nombre': ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                          'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][mes],
            'monto': monto_mes,
            'cantidad': cantidad_mes
        })
    
    # Top 10 clientes por monto pagado
    top_clientes = pagos.filter(estado='pagado').values(
        'cliente__nombre', 'cliente__apellido1', 'cliente__id'
    ).annotate(
        total_pagado=Sum('monto'),
        cantidad_pagos=Count('id')
    ).order_by('-total_pagado')[:10]
    
    # Clientes morosos (con pagos vencidos)
    clientes_morosos = pagos.filter(estado='vencido').values(
        'cliente__nombre', 'cliente__apellido1', 'cliente__id'
    ).annotate(
        total_vencido=Sum('monto'),
        cantidad_vencidos=Count('id')
    ).order_by('-total_vencido')[:10]
    
    # Métodos de pago más usados
    metodos_pago = pagos.filter(estado='pagado', metodo_pago__isnull=False).values(
        'metodo_pago'
    ).annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    # Promedio de pago
    promedio_pago = pagos.filter(estado='pagado').aggregate(Avg('monto'))['monto__avg'] or 0
    
    context = {
        'anio': anio,
        'años_disponibles': años_disponibles,
        'total_pagos': total_pagos,
        'total_monto': total_monto,
        'monto_pagado': monto_pagado,
        'monto_pendiente': monto_pendiente,
        'ingresos_por_mes': ingresos_por_mes,
        'top_clientes': top_clientes,
        'clientes_morosos': clientes_morosos,
        'metodos_pago': metodos_pago,
        'promedio_pago': promedio_pago,
    }
    
    return render(request, 'pagos/pago_reportes.html', context)


def pago_procesar_online(request, pk):
    """Inicia el proceso de pago en línea para un pago."""
    pago = get_object_or_404(Pago, pk=pk)
    
    # Verificar permisos: solo el cliente dueño o staff puede pagar
    from clientes.portal_views import es_cliente, obtener_cliente_desde_usuario
    if request.user.is_authenticated:
        if es_cliente(request.user):
            cliente = obtener_cliente_desde_usuario(request.user)
            # Verificar que el pago pertenezca al cliente
            if pago.cliente != cliente:
                messages.error(request, 'No tienes permiso para acceder a este pago.')
                return redirect('clientes:portal_mis_pagos')
            # Verificar que si el pago tiene instalación, esa instalación pertenezca al cliente
            if pago.instalacion and pago.instalacion.cliente != cliente:
                messages.error(request, 'No tienes permiso para acceder a este pago.')
                return redirect('clientes:portal_mis_pagos')
        else:
            # Si no es cliente, verificar que tenga permiso para ver pagos
            from core.roles_utils import usuario_tiene_permiso
            if not usuario_tiene_permiso(request.user, 'ver_pagos'):
                messages.error(request, 'No tienes permiso para acceder a esta página.')
                return redirect('login')
    else:
        messages.error(request, 'Debes iniciar sesión para realizar un pago.')
        return redirect('clientes:portal_login')
    
    # Verificar que el pago esté pendiente o vencido (no pagado)
    if pago.estado == 'pagado':
        messages.warning(request, 'Este pago ya ha sido pagado. No se puede procesar nuevamente.')
        if es_cliente(request.user):
            return redirect('clientes:portal_detalle_pago', pago_id=pk)
        return redirect('pagos:pago_detail', pk=pk)
    
    if pago.estado == 'cancelado':
        messages.warning(request, 'Este pago ha sido cancelado. No se puede procesar.')
        if es_cliente(request.user):
            return redirect('clientes:portal_detalle_pago', pago_id=pk)
        return redirect('pagos:pago_detail', pk=pk)
    
    # Verificar si ya existe una transacción completada para este pago
    transaccion_completada = pago.transacciones.filter(estado='completada').exists()
    if transaccion_completada:
        messages.warning(request, 'Este pago ya tiene una transacción completada. No se puede procesar nuevamente.')
        # Actualizar el estado del pago si no está marcado como pagado
        if pago.estado != 'pagado':
            pago.marcar_como_pagado(metodo_pago='tarjeta')
        if es_cliente(request.user):
            return redirect('clientes:portal_detalle_pago', pago_id=pk)
        return redirect('pagos:pago_detail', pk=pk)
    
    # Verificar si hay una transacción pendiente reciente (últimos 5 minutos)
    # para evitar múltiples intentos simultáneos
    from datetime import timedelta
    transaccion_reciente = pago.transacciones.filter(
        estado='pendiente',
        fecha_creacion__gte=timezone.now() - timedelta(minutes=5)
    ).exists()
    if transaccion_reciente:
        messages.info(request, 'Ya existe un proceso de pago en curso para este pago. Por favor espera unos momentos.')
        if es_cliente(request.user):
            return redirect('clientes:portal_detalle_pago', pago_id=pk)
        return redirect('pagos:pago_detail', pk=pk)
    
    # Si es POST, procesar la selección de pasarela
    if request.method == 'POST':
        pasarela = request.POST.get('pasarela', 'stripe')
        
        # Obtener URLs de retorno
        # NOTA: Mercado Pago NO acepta placeholders ({payment_id}) en back_urls
        # Las URLs deben ser completas. Mercado Pago agregará los parámetros automáticamente.
        base_url = getattr(settings, 'SITE_URL', None) or request.build_absolute_uri('/').rstrip('/')
        
        # Limpiar base_url de comentarios o espacios si vienen del .env
        if '#' in base_url:
            base_url = base_url.split('#')[0]
        base_url = base_url.strip().rstrip('/')
        
        # URLs sin placeholders - Mercado Pago agregará payment_id automáticamente en la redirección
        return_url = f"{base_url}/pagos/{pk}/pago-exitoso/"
        cancel_url = f"{base_url}/pagos/{pk}/pago-cancelado/"
        
        try:
            # Crear intento de pago con la pasarela seleccionada
            gateway = PaymentGateway(pasarela=pasarela)
            resultado = gateway.crear_intento_pago(pago, return_url, cancel_url)
            
            if resultado.get('success'):
                # Redirigir a la pasarela de pago
                return redirect(resultado['url'])
            else:
                messages.error(request, f'Error al procesar el pago: {resultado.get("error", "Error desconocido")}')
                if es_cliente(request.user):
                    return redirect('clientes:portal_detalle_pago', pago_id=pk)
                return redirect('pagos:pago_detail', pk=pk)
        except ImportError as e:
            messages.error(request, f'La pasarela seleccionada no está disponible: {str(e)}')
            if es_cliente(request.user):
                return redirect('clientes:portal_detalle_pago', pago_id=pk)
            return redirect('pagos:pago_detail', pk=pk)
        except Exception as e:
            error_msg = str(e)
            logger.error(f"Error inesperado al procesar pago {pk}: {error_msg}", exc_info=True)
            
            # Mensaje más amigable según el tipo de error
            if "no está instalado" in error_msg:
                messages.error(request, f'La pasarela seleccionada no está disponible: {error_msg}')
            elif "no está configurada" in error_msg or "no configurada" in error_msg:
                messages.error(request, f'Error de configuración: {error_msg}. Verifica tus credenciales en .env')
            else:
                messages.error(request, f'Error al procesar el pago: {error_msg}')
            
            if es_cliente(request.user):
                return redirect('clientes:portal_detalle_pago', pago_id=pk)
            return redirect('pagos:pago_detail', pk=pk)
    
    # Si es GET, mostrar formulario de selección de pasarela
    pasarelas_disponibles = []
    
    # Verificar Stripe
    try:
        from pagos.payment_gateway import STRIPE_AVAILABLE
        if STRIPE_AVAILABLE and getattr(settings, 'STRIPE_SECRET_KEY', ''):
            pasarelas_disponibles.append(('stripe', 'Stripe', 'Tarjetas de crédito y débito'))
    except:
        pass
    
    # Verificar Mercado Pago
    try:
        from pagos.payment_gateway import MERCADOPAGO_AVAILABLE
        if MERCADOPAGO_AVAILABLE and getattr(settings, 'MERCADOPAGO_ACCESS_TOKEN', ''):
            pasarelas_disponibles.append(('mercadopago', 'Mercado Pago', 'Tarjetas, efectivo y más'))
    except:
        pass
    
    # Verificar PayPal
    try:
        if getattr(settings, 'PAYPAL_CLIENT_ID', '') and getattr(settings, 'PAYPAL_SECRET', ''):
            pasarelas_disponibles.append(('paypal', 'PayPal', 'PayPal y tarjetas'))
    except:
        pass
    
    if not pasarelas_disponibles:
        messages.warning(request, 'No hay pasarelas de pago configuradas.')
        if es_cliente(request.user):
            return redirect('clientes:portal_detalle_pago', pago_id=pk)
        return redirect('pagos:pago_detail', pk=pk)
    
    # Detectar si estamos en localhost para mostrar advertencia
    base_url = getattr(settings, 'SITE_URL', None) or request.build_absolute_uri('/').rstrip('/')
    es_localhost = any(host in base_url.lower() for host in ['localhost', '127.0.0.1', '0.0.0.0'])
    
    context = {
        'pago': pago,
        'pasarelas_disponibles': pasarelas_disponibles,
        'es_localhost': es_localhost,
    }
    
    return render(request, 'pagos/pago_seleccionar_pasarela.html', context)


def pago_exitoso(request, pk):
    """Vista que se muestra después de un pago exitoso."""
    pago = get_object_or_404(Pago, pk=pk)
    
    # Verificar permisos
    from clientes.portal_views import es_cliente, obtener_cliente_desde_usuario
    if request.user.is_authenticated:
        if es_cliente(request.user):
            cliente = obtener_cliente_desde_usuario(request.user)
            # Verificar que el pago pertenezca al cliente
            if pago.cliente != cliente:
                messages.error(request, 'No tienes permiso para acceder a este pago.')
                return redirect('clientes:portal_mis_pagos')
            # Verificar que si el pago tiene instalación, esa instalación pertenezca al cliente
            if pago.instalacion and pago.instalacion.cliente != cliente:
                messages.error(request, 'No tienes permiso para acceder a este pago.')
                return redirect('clientes:portal_mis_pagos')
        else:
            # Si no es cliente, verificar que tenga permiso para ver pagos
            from core.roles_utils import usuario_tiene_permiso
            if not usuario_tiene_permiso(request.user, 'ver_pagos'):
                messages.error(request, 'No tienes permiso para acceder a esta página.')
                return redirect('login')
    
    # Verificar si el pago ya está pagado (evitar procesamiento duplicado)
    if pago.estado == 'pagado':
        # Verificar si hay una transacción completada
        transaccion_completada = pago.transacciones.filter(estado='completada').first()
        if transaccion_completada:
            # Si es cliente, redirigir directamente a mis-pagos
            if request.user.is_authenticated and es_cliente(request.user):
                messages.info(request, 'Este pago ya fue procesado exitosamente anteriormente.')
                # Redirigir específicamente a localhost:8000
                from django.http import HttpResponseRedirect
                return HttpResponseRedirect('http://localhost:8000/clientes/portal/mis-pagos/')
            # Si es staff, mostrar la página de éxito
            messages.success(request, 'Este pago ya fue procesado exitosamente anteriormente.')
            transacciones = pago.transacciones.all().order_by('-fecha_creacion')
            context = {
                'pago': pago,
                'transacciones': transacciones,
            }
            return render(request, 'pagos/pago_exitoso.html', context)
    
    # Verificar según la pasarela
    session_id = request.GET.get('session_id')  # Stripe
    payment_id = request.GET.get('payment_id') or request.GET.get('paymentId')  # Mercado Pago (puede venir de diferentes formas)
    # PayPal puede usar 'token' o 'paymentId' dependiendo de la versión de la API
    paypal_token = request.GET.get('token') or request.GET.get('paymentId')  # PayPal
    
    # Log para debugging
    logger.info(f"pago_exitoso llamado para pago {pk}. GET params: {dict(request.GET)}")
    logger.info(f"payment_id encontrado: {payment_id}, session_id: {session_id}, paypal_token: {paypal_token}")
    
    if session_id:
        # Verificar el pago con Stripe
        gateway = PaymentGateway(pasarela='stripe')
        resultado = gateway.verificar_pago(session_id)
        
        if resultado.get('success') and resultado.get('estado') == 'completada':
            messages.success(request, '¡Pago procesado exitosamente!')
            pago.refresh_from_db()
        elif resultado.get('success'):
            messages.info(request, 'El pago está siendo procesado. Recibirás una confirmación pronto.')
        else:
            messages.warning(request, f'No se pudo verificar el pago: {resultado.get("error", "Error desconocido")}')
    elif payment_id:
        # Mercado Pago - verificar y actualizar transacción
        try:
            # Buscar la transacción por payment_id
            transaccion = TransaccionPago.objects.filter(
                id_transaccion_pasarela=payment_id,
                pasarela='mercadopago'
            ).first()
            
            # Si no encontramos por payment_id, buscar por preference_id o external_reference
            if not transaccion:
                # Buscar por el external_reference que debería ser el ID del pago
                gateway = PaymentGateway(pasarela='mercadopago')
                try:
                    payment_response = gateway.mp.payment().get(int(payment_id))
                    if payment_response.get("status") == 200:
                        payment = payment_response.get("response", {})
                        external_reference = payment.get("external_reference")
                        if external_reference:
                            # Buscar transacción por el pago asociado
                            try:
                                pago_id = int(external_reference)
                                transaccion = TransaccionPago.objects.filter(
                                    pago__id=pago_id,
                                    pasarela='mercadopago'
                                ).first()
                            except (ValueError, TypeError):
                                pass
                except Exception as e:
                    logger.warning(f"Error al buscar transacción por payment_id {payment_id}: {str(e)}")
            
            # Si todavía no encontramos, buscar cualquier transacción pendiente de este pago
            if not transaccion:
                transaccion = TransaccionPago.objects.filter(
                    pago=pago,
                    pasarela='mercadopago',
                    estado__in=['pendiente', 'completada']
                ).order_by('-fecha_creacion').first()
            
            if transaccion:
                # Verificar el estado del pago con la API de Mercado Pago
                gateway = PaymentGateway(pasarela='mercadopago')
                try:
                    payment_response = gateway.mp.payment().get(int(payment_id))
                    
                    if payment_response.get("status") == 200:
                        payment = payment_response.get("response", {})
                        payment_status = payment.get("status", "")
                        
                        # Actualizar datos de la transacción
                        if not transaccion.id_transaccion_pasarela or transaccion.id_transaccion_pasarela != payment_id:
                            transaccion.id_transaccion_pasarela = payment_id
                        
                        if not transaccion.datos_respuesta:
                            transaccion.datos_respuesta = {}
                        transaccion.datos_respuesta['payment'] = payment
                        
                        # Si el pago fue aprobado, marcar como completada
                        if payment_status == "approved":
                            # Verificar que el pago no esté ya pagado (evitar duplicidad)
                            if pago.estado == 'pagado':
                                # Verificar si esta transacción ya está completada
                                if transaccion.estado == 'completada':
                                    messages.success(request, 'Este pago ya fue procesado exitosamente anteriormente.')
                                    logger.info(f"Intento de procesar pago ya pagado (Pago ID: {pago.id}, Payment ID: {payment_id})")
                                else:
                                    # Actualizar solo la transacción, no el pago
                                    transaccion.estado = 'completada'
                                    transaccion.fecha_completada = timezone.now()
                                    transaccion.save()
                                    messages.success(request, '¡Pago procesado exitosamente!')
                            elif transaccion.estado != 'completada':
                                transaccion.marcar_como_completada()
                                messages.success(request, '¡Pago procesado exitosamente!')
                                logger.info(f"Pago {pago.id} marcado como pagado por Mercado Pago payment_id: {payment_id}")
                            else:
                                messages.success(request, '¡Pago procesado exitosamente!')
                        elif payment_status == "pending":
                            transaccion.estado = 'pendiente'
                            transaccion.save()
                            messages.info(request, 'El pago está siendo procesado. Recibirás una confirmación pronto.')
                        elif payment_status == "rejected":
                            transaccion.estado = 'fallida'
                            transaccion.mensaje_error = payment.get("status_detail", "Pago rechazado")
                            transaccion.save()
                            messages.warning(request, f'El pago fue rechazado: {transaccion.mensaje_error}')
                        elif payment_status == "cancelled":
                            transaccion.estado = 'fallida'
                            transaccion.mensaje_error = "Pago cancelado"
                            transaccion.save()
                            messages.warning(request, 'El pago fue cancelado.')
                        else:
                            # Estado desconocido
                            transaccion.estado = 'pendiente'
                            transaccion.save()
                            messages.info(request, f'Estado del pago: {payment_status}. Estamos verificando el pago...')
                    else:
                        # Error al obtener el pago de Mercado Pago
                        error_msg = payment_response.get("message", "Error desconocido")
                        logger.error(f"Error al obtener pago de Mercado Pago {payment_id}: {error_msg}")
                        messages.warning(request, 'No se pudo verificar el estado del pago con Mercado Pago.')
                except Exception as e:
                    logger.error(f"Error al verificar pago de Mercado Pago {payment_id}: {str(e)}", exc_info=True)
                    # Si la transacción ya está completada, asumir que el pago fue exitoso
                    if transaccion.estado == 'completada':
                        messages.success(request, '¡Pago procesado exitosamente!')
                    else:
                        messages.info(request, 'El pago está siendo procesado. Recibirás una confirmación pronto.')
            else:
                # Transacción no encontrada, pero el payment_id existe - crear una nueva transacción
                gateway = PaymentGateway(pasarela='mercadopago')
                try:
                    payment_response = gateway.mp.payment().get(int(payment_id))
                    if payment_response.get("status") == 200:
                        payment = payment_response.get("response", {})
                        payment_status = payment.get("status", "")
                        
                        # Crear nueva transacción
                        transaccion = TransaccionPago.objects.create(
                            pago=pago,
                            pasarela='mercadopago',
                            estado='pendiente',
                            id_transaccion_pasarela=payment_id,
                            monto=pago.monto,
                            moneda='MXN',
                            datos_respuesta={'payment': payment}
                        )
                        
                        # Si el pago fue aprobado, marcar como completada
                        if payment_status == "approved":
                            transaccion.marcar_como_completada()
                            messages.success(request, '¡Pago procesado exitosamente!')
                            logger.info(f"Pago {pago.id} marcado como pagado por Mercado Pago payment_id: {payment_id}")
                        elif payment_status == "pending":
                            messages.info(request, 'El pago está siendo procesado. Recibirás una confirmación pronto.')
                        else:
                            messages.info(request, f'Estado del pago: {payment_status}')
                except Exception as e:
                    logger.error(f"Error al crear transacción para payment_id {payment_id}: {str(e)}", exc_info=True)
                    messages.warning(request, 'No se encontró la transacción y no se pudo crear una nueva.')
            
            # Refrescar el pago desde la base de datos para obtener el estado actualizado
            pago.refresh_from_db()
            
            # Verificar si el pago se marcó como pagado
            if pago.estado != 'pagado':
                # Si hay una transacción completada pero el pago no está marcado como pagado, actualizar
                transaccion_completada = pago.transacciones.filter(estado='completada').first()
                if transaccion_completada:
                    logger.warning(f"Pago {pago.id} tiene transacción completada pero estado no es 'pagado'. Actualizando...")
                    # Forzar actualización del estado del pago
                    pago.marcar_como_pagado(
                        metodo_pago='tarjeta',
                        referencia=transaccion_completada.id_transaccion_pasarela or f"MP-{transaccion_completada.id}"
                    )
                    pago.refresh_from_db()
                    logger.info(f"Pago {pago.id} actualizado a estado 'pagado'")
        except TransaccionPago.DoesNotExist:
            messages.warning(request, 'No se encontró la transacción.')
        except ValueError:
            messages.warning(request, 'ID de pago inválido.')
        except Exception as e:
            logger.error(f"Error inesperado al procesar pago de Mercado Pago: {str(e)}", exc_info=True)
            messages.warning(request, f'Error al procesar el pago: {str(e)}')
    elif paypal_token:
        # PayPal - verificar y capturar el pago
        # paypal_token es el token (order_id) que PayPal devuelve
        try:
            gateway = PaymentGateway(pasarela='paypal')
            # Buscar la transacción por order_id
            transaccion = TransaccionPago.objects.filter(
                id_transaccion_pasarela=paypal_token,
                pasarela='paypal'
            ).first()
            
            if not transaccion:
                # Si no encontramos la transacción, intentar buscar por cualquier orden de PayPal para este pago
                transaccion = TransaccionPago.objects.filter(
                    pago=pago,
                    pasarela='paypal',
                    estado='pendiente'
                ).first()
            
            if transaccion:
                # Verificar el estado de la orden primero
                access_token = gateway._obtener_paypal_access_token()
                if access_token:
                    api_url = 'https://api-m.sandbox.paypal.com' if gateway.paypal_mode == 'sandbox' else 'https://api-m.paypal.com'
                    headers = {
                        'Content-Type': 'application/json',
                        'Authorization': f'Bearer {access_token}'
                    }
                    
                    # Obtener detalles de la orden
                    order_response = requests.get(
                        f'{api_url}/v2/checkout/orders/{paypal_token}',
                        headers=headers
                    )
                    
                    if order_response.status_code == 200:
                        order = order_response.json()
                        order_status = order.get('status', '')
                        
                        # Si la orden está aprobada pero no capturada, capturarla
                        if order_status == 'APPROVED':
                            # Capturar el pago en PayPal
                            capture_response = requests.post(
                                f'{api_url}/v2/checkout/orders/{paypal_token}/capture',
                                headers=headers,
                                json={}
                            )
                            
                            if capture_response.status_code == 201:
                                capture_data = capture_response.json()
                                # Actualizar transacción con el order_id correcto si no lo tenía
                                if transaccion.id_transaccion_pasarela != paypal_token:
                                    transaccion.id_transaccion_pasarela = paypal_token
                                
                                transaccion.estado = 'completada'
                                transaccion.fecha_completada = timezone.now()
                                if not transaccion.datos_respuesta:
                                    transaccion.datos_respuesta = {}
                                transaccion.datos_respuesta['order'] = order
                                transaccion.datos_respuesta['capture'] = capture_data
                                transaccion.save()
                                transaccion.marcar_como_completada()
                                messages.success(request, '¡Pago procesado exitosamente!')
                            else:
                                error_data = capture_response.json() if capture_response.content else {}
                                error_msg = error_data.get('message', 'Error al capturar el pago')
                                logger.error(f"Error al capturar pago de PayPal: {error_msg}")
                                messages.warning(request, f'Error al procesar el pago: {error_msg}')
                        elif order_status == 'COMPLETED':
                            # La orden ya está completada
                            if transaccion.estado != 'completada':
                                transaccion.estado = 'completada'
                                transaccion.fecha_completada = timezone.now()
                                if not transaccion.datos_respuesta:
                                    transaccion.datos_respuesta = {}
                                transaccion.datos_respuesta['order'] = order
                                transaccion.save()
                                transaccion.marcar_como_completada()
                            messages.success(request, '¡Pago procesado exitosamente!')
                        elif order_status == 'CREATED':
                            messages.info(request, 'El pago está pendiente de aprobación.')
                        else:
                            messages.warning(request, f'Estado del pago: {order_status}')
                    else:
                        error_data = order_response.json() if order_response.content else {}
                        error_msg = error_data.get('message', 'Error al verificar el pago')
                        logger.error(f"Error al verificar orden de PayPal: {error_msg}")
                        messages.warning(request, f'Error al verificar el pago: {error_msg}')
                else:
                    messages.warning(request, 'No se pudo obtener el token de acceso de PayPal.')
            else:
                messages.warning(request, 'No se encontró la transacción asociada a este pago.')
            pago.refresh_from_db()
        except Exception as e:
            logger.error(f"Error al procesar pago de PayPal: {str(e)}", exc_info=True)
            messages.warning(request, f'Error al verificar el pago: {str(e)}')
    else:
        # Si no hay payment_id, session_id ni paypal_token, verificar si hay transacciones completadas recientes
        # Esto puede pasar si el usuario hace clic en "Volver al sitio" sin parámetros
        logger.info(f"No se encontró payment_id, session_id ni paypal_token. Verificando transacciones completadas para pago {pk}")
        transaccion_completada = pago.transacciones.filter(
            estado='completada',
            pasarela='mercadopago'
        ).order_by('-fecha_completada').first()
        
        if transaccion_completada:
            # Verificar que el pago esté marcado como pagado
            if pago.estado != 'pagado':
                logger.warning(f"Pago {pago.id} tiene transacción completada reciente pero estado no es 'pagado'. Actualizando...")
                pago.marcar_como_pagado(
                    metodo_pago='tarjeta',
                    referencia=transaccion_completada.id_transaccion_pasarela or f"MP-{transaccion_completada.id}"
                )
                pago.refresh_from_db()
                logger.info(f"Pago {pago.id} actualizado a estado 'pagado' basado en transacción completada")
            messages.success(request, '¡Pago procesado exitosamente!')
        else:
            # Buscar transacciones pendientes recientes (últimas 10 minutos) que puedan haberse completado
            from datetime import timedelta
            transacciones_recientes = pago.transacciones.filter(
                pasarela='mercadopago',
                fecha_creacion__gte=timezone.now() - timedelta(minutes=10)
            ).order_by('-fecha_creacion')
            
            if transacciones_recientes.exists():
                # Intentar verificar con Mercado Pago si alguna se completó
                gateway = PaymentGateway(pasarela='mercadopago')
                for trans in transacciones_recientes:
                    if trans.id_transaccion_pasarela:
                        try:
                            payment_response = gateway.mp.payment().get(int(trans.id_transaccion_pasarela))
                            if payment_response.get("status") == 200:
                                payment = payment_response.get("response", {})
                                payment_status = payment.get("status", "")
                                
                                if payment_status == "approved" and trans.estado != 'completada':
                                    trans.marcar_como_completada()
                                    pago.refresh_from_db()
                                    messages.success(request, '¡Pago procesado exitosamente!')
                                    logger.info(f"Pago {pago.id} verificado y marcado como pagado desde transacción {trans.id}")
                                    break
                        except Exception as e:
                            logger.warning(f"Error al verificar transacción {trans.id}: {str(e)}")
                            continue
            
            if pago.estado != 'pagado':
                messages.info(request, 'El pago está siendo verificado. Si ya realizaste el pago, el estado se actualizará automáticamente.')
    
    # Refrescar el pago desde la base de datos antes de renderizar para asegurar el estado más actualizado
    pago.refresh_from_db()
    
    # Verificación final: si hay una transacción completada pero el pago no está marcado como pagado, actualizar
    if pago.estado != 'pagado':
        transaccion_completada = pago.transacciones.filter(estado='completada').first()
        if transaccion_completada:
            logger.warning(f"Verificación final: Pago {pago.id} tiene transacción completada pero estado no es 'pagado'. Actualizando...")
            pago.marcar_como_pagado(
                metodo_pago='tarjeta',
                referencia=transaccion_completada.id_transaccion_pasarela or f"MP-{transaccion_completada.id}"
            )
            pago.refresh_from_db()
            logger.info(f"Pago {pago.id} actualizado a estado 'pagado' en verificación final")
    
    # Si es un cliente, redirigir directamente a mis-pagos después de procesar el pago
    if request.user.is_authenticated and es_cliente(request.user):
        messages.success(request, f'¡Pago procesado exitosamente! El pago de ${pago.monto} ha sido registrado.')
        # Redirigir específicamente a localhost:8000
        # Nota: Si vienes desde ngrok, necesitarás hacer clic en "Visit Site" la primera vez
        from django.http import HttpResponseRedirect
        return HttpResponseRedirect('http://localhost:8000/clientes/portal/mis-pagos/')
    
    # Obtener transacciones relacionadas
    transacciones = pago.transacciones.all().order_by('-fecha_creacion')
    
    # Determinar la URL de redirección según el tipo de usuario
    from core.roles_utils import usuario_tiene_permiso
    redirect_url = None
    if request.user.is_authenticated:
        if usuario_tiene_permiso(request.user, 'ver_pagos'):
            # Redirigir al detalle del pago en el admin
            redirect_url = 'pagos:pago_detail'
    
    context = {
        'pago': pago,
        'transacciones': transacciones,
        'redirect_url': redirect_url,
        'es_cliente': False,  # Solo llegamos aquí si NO es cliente (ya redirigimos arriba)
    }
    
    return render(request, 'pagos/pago_exitoso.html', context)


def pago_cancelado(request, pk):
    """Vista que se muestra cuando se cancela un pago."""
    pago = get_object_or_404(Pago, pk=pk)
    messages.info(request, 'El pago fue cancelado. Puedes intentar nuevamente cuando lo desees.')
    
    # Redirigir según el tipo de usuario
    from clientes.portal_views import es_cliente
    if request.user.is_authenticated and es_cliente(request.user):
        return redirect('clientes:portal_detalle_pago', pago_id=pk)
    return redirect('pagos:pago_detail', pk=pk)


def stripe_webhook(request):
    """Endpoint para recibir webhooks de Stripe."""
    payload = request.body
    sig_header = request.META.get('HTTP_STRIPE_SIGNATURE')
    
    gateway = PaymentGateway(pasarela='stripe')
    resultado = gateway.procesar_webhook(payload, sig_header)
    
    if resultado.get('success'):
        return JsonResponse({'status': 'success'}, status=200)
    else:
        return JsonResponse({'status': 'error', 'message': resultado.get('error')}, status=400)


@csrf_exempt
def mercadopago_webhook(request):
    """Endpoint para recibir webhooks de Mercado Pago."""
    import json
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            tipo = data.get('type')
            datos = data.get('data', {})
            
            logger.info(f"Webhook de Mercado Pago recibido: tipo={tipo}, datos={datos}")
            
            if tipo == 'payment':
                payment_id = datos.get('id')
                
                if not payment_id:
                    logger.warning("Webhook de Mercado Pago sin payment_id")
                    return JsonResponse({'status': 'error', 'message': 'payment_id requerido'}, status=400)
                
                # Buscar la transacción
                transaccion = TransaccionPago.objects.filter(
                    id_transaccion_pasarela=payment_id,
                    pasarela='mercadopago'
                ).first()
                
                # Si no encontramos por payment_id, buscar por external_reference
                if not transaccion:
                    try:
                        gateway = PaymentGateway(pasarela='mercadopago')
                        payment_response = gateway.mp.payment().get(int(payment_id))
                        
                        if payment_response.get("status") == 200:
                            payment = payment_response.get("response", {})
                            external_reference = payment.get("external_reference")
                            
                            if external_reference:
                                try:
                                    pago_id = int(external_reference)
                                    # Buscar transacción por el pago asociado
                                    transaccion = TransaccionPago.objects.filter(
                                        pago__id=pago_id,
                                        pasarela='mercadopago'
                                    ).order_by('-fecha_creacion').first()
                                    
                                    # Si encontramos la transacción, actualizar el payment_id
                                    if transaccion and not transaccion.id_transaccion_pasarela:
                                        transaccion.id_transaccion_pasarela = payment_id
                                        transaccion.save()
                                except (ValueError, TypeError):
                                    pass
                    except Exception as e:
                        logger.warning(f"Error al buscar transacción por external_reference: {str(e)}")
                
                if transaccion:
                    # Obtener detalles actualizados del pago
                    try:
                        gateway = PaymentGateway(pasarela='mercadopago')
                        payment_response = gateway.mp.payment().get(int(payment_id))
                        
                        if payment_response.get("status") == 200:
                            payment = payment_response.get("response", {})
                            payment_status = payment.get("status", "")
                            
                            # Actualizar datos de la transacción
                            if not transaccion.id_transaccion_pasarela:
                                transaccion.id_transaccion_pasarela = payment_id
                            
                            if not transaccion.datos_respuesta:
                                transaccion.datos_respuesta = {}
                            transaccion.datos_respuesta['payment'] = payment
                            
                            # Actualizar estado según el status del pago
                            if payment_status == "approved":
                                # Verificar que el pago no esté ya pagado (evitar duplicidad)
                                if transaccion.pago.estado == 'pagado':
                                    # Solo actualizar la transacción si no está completada
                                    if transaccion.estado != 'completada':
                                        transaccion.estado = 'completada'
                                        transaccion.fecha_completada = timezone.now()
                                        transaccion.save()
                                    logger.warning(f"Webhook: Intento de procesar pago ya pagado (Pago ID: {transaccion.pago.id}, Payment ID: {payment_id})")
                                elif transaccion.estado != 'completada':
                                    transaccion.marcar_como_completada()
                                    logger.info(f"Webhook: Pago {transaccion.pago.id} marcado como pagado por Mercado Pago payment_id: {payment_id}")
                            elif payment_status == "rejected":
                                transaccion.estado = 'fallida'
                                transaccion.mensaje_error = payment.get("status_detail", "Pago rechazado")
                                transaccion.save()
                                logger.info(f"Webhook: Pago {transaccion.pago.id} marcado como rechazado")
                            elif payment_status == "cancelled":
                                transaccion.estado = 'fallida'
                                transaccion.mensaje_error = "Pago cancelado"
                                transaccion.save()
                                logger.info(f"Webhook: Pago {transaccion.pago.id} marcado como cancelado")
                            elif payment_status == "pending":
                                transaccion.estado = 'pendiente'
                                transaccion.save()
                                logger.info(f"Webhook: Pago {transaccion.pago.id} sigue pendiente")
                        else:
                            error_msg = payment_response.get("message", "Error desconocido")
                            logger.error(f"Error al obtener pago de Mercado Pago {payment_id}: {error_msg}")
                    except Exception as e:
                        logger.error(f"Error al procesar webhook de Mercado Pago para payment_id {payment_id}: {str(e)}", exc_info=True)
                else:
                    logger.warning(f"Webhook: Transacción no encontrada para payment_id: {payment_id}")
            
            return JsonResponse({'status': 'success'}, status=200)
        except json.JSONDecodeError as e:
            logger.error(f"Error al parsear JSON del webhook de Mercado Pago: {str(e)}")
            return JsonResponse({'status': 'error', 'message': 'JSON inválido'}, status=400)
        except Exception as e:
            logger.error(f"Error al procesar webhook de Mercado Pago: {str(e)}", exc_info=True)
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)


@login_required
def pago_reembolsar(request, transaccion_id):
    """Procesa un reembolso de una transacción."""
    transaccion = get_object_or_404(TransaccionPago, pk=transaccion_id)
    
    # Verificar que la transacción esté completada
    if transaccion.estado != 'completada':
        messages.error(request, 'Solo se pueden reembolsar transacciones completadas.')
        return redirect('pagos:pago_detail', pk=transaccion.pago.pk)
    
    # Verificar que no esté ya reembolsada
    if transaccion.estado == 'reembolsada':
        messages.warning(request, 'Esta transacción ya ha sido reembolsada.')
        return redirect('pagos:pago_detail', pk=transaccion.pago.pk)
    
    if request.method == 'POST':
        monto_parcial = request.POST.get('monto_parcial')
        motivo = request.POST.get('motivo', '')
        
        try:
            monto_parcial = float(monto_parcial) if monto_parcial else None
            
            # Validar monto parcial
            if monto_parcial:
                if monto_parcial <= 0:
                    messages.error(request, 'El monto del reembolso debe ser mayor a cero.')
                    return redirect('pagos:pago_reembolsar', transaccion_id=transaccion_id)
                if monto_parcial > float(transaccion.monto):
                    messages.error(request, 'El monto del reembolso no puede ser mayor al monto de la transacción.')
                    return redirect('pagos:pago_reembolsar', transaccion_id=transaccion_id)
            
            # Procesar reembolso
            gateway = PaymentGateway(pasarela=transaccion.pasarela)
            resultado = gateway.procesar_reembolso(transaccion, monto_parcial, motivo)
            
            if resultado.get('success'):
                messages.success(
                    request,
                    f'Reembolso procesado exitosamente. Monto: ${resultado.get("amount", 0):.2f}'
                )
                return redirect('pagos:pago_detail', pk=transaccion.pago.pk)
            else:
                messages.error(request, f'Error al procesar el reembolso: {resultado.get("error", "Error desconocido")}')
                return redirect('pagos:pago_reembolsar', transaccion_id=transaccion_id)
        except ValueError:
            messages.error(request, 'El monto del reembolso debe ser un número válido.')
            return redirect('pagos:pago_reembolsar', transaccion_id=transaccion_id)
        except Exception as e:
            logger.error(f"Error al procesar reembolso: {str(e)}")
            messages.error(request, f'Error inesperado: {str(e)}')
            return redirect('pagos:pago_reembolsar', transaccion_id=transaccion_id)
    
    context = {
        'transaccion': transaccion,
        'pago': transaccion.pago,
    }
    
    return render(request, 'pagos/pago_reembolsar.html', context)
