from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from datetime import date
from .models import Cliente
from .forms import ClienteForm


@login_required
def cliente_list(request):
    """Lista todos los clientes con búsqueda y paginación."""
    # Filtro para mostrar eliminados
    mostrar_eliminados = request.GET.get('mostrar_eliminados', '') == '1'
    
    if mostrar_eliminados:
        clientes = Cliente.all_objects.all()
    else:
        clientes = Cliente.objects.all()  # Usa el manager personalizado que filtra eliminados
    
    # Búsqueda simple (compatibilidad)
    query = request.GET.get('q', '')
    if query:
        clientes = clientes.filter(
            Q(nombre__icontains=query) |
            Q(apellido1__icontains=query) |
            Q(apellido2__icontains=query) |
            Q(telefono__icontains=query) |
            Q(email__icontains=query) |
            Q(ciudad__icontains=query)
        )
    
    # Búsqueda avanzada con múltiples criterios simultáneos
    nombre = request.GET.get('nombre', '')
    if nombre:
        clientes = clientes.filter(Q(nombre__icontains=nombre) | Q(apellido1__icontains=nombre) | Q(apellido2__icontains=nombre))
    
    telefono = request.GET.get('telefono', '')
    if telefono:
        clientes = clientes.filter(telefono__icontains=telefono)
    
    email = request.GET.get('email', '')
    if email:
        clientes = clientes.filter(email__icontains=email)
    
    ciudad = request.GET.get('ciudad', '')
    if ciudad:
        clientes = clientes.filter(ciudad__icontains=ciudad)
    
    # Filtro por estado
    estado_filter = request.GET.get('estado', '')
    if estado_filter:
        clientes = clientes.filter(estado_cliente=estado_filter)
    
    # Ordenamiento
    orden = request.GET.get('orden', '-created_at')
    # Si el orden es por nombre, ordenar por apellido1 y luego nombre
    if orden in ['nombre', '-nombre']:
        if orden == 'nombre':
            clientes = clientes.order_by('apellido1', 'apellido2', 'nombre')
        else:
            clientes = clientes.order_by('-apellido1', '-apellido2', '-nombre')
    else:
        clientes = clientes.order_by(orden)
    
    # Paginación
    paginator = Paginator(clientes, 15)  # 15 clientes por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'query': query,
        'nombre': nombre,
        'telefono': telefono,
        'email': email,
        'ciudad': ciudad,
        'estado_filter': estado_filter,
        'orden': orden,
        'estados': Cliente.ESTADO_CHOICES,
        'mostrar_eliminados': mostrar_eliminados,
    }
    
    return render(request, 'clientes/cliente_list.html', context)


@login_required
def cliente_detail(request, pk):
    """Muestra el detalle de un cliente."""
    cliente = get_object_or_404(Cliente.all_objects, pk=pk)  # Usar all_objects para permitir ver eliminados
    
    # Limpiar credenciales si se solicita
    if request.GET.get('limpiar_credenciales') == '1' and 'credenciales_cliente' in request.session:
        del request.session['credenciales_cliente']
        request.session.modified = True
    
    # Obtener instalaciones relacionadas
    instalaciones = cliente.instalaciones.all()[:5]
    
    # Obtener pagos relacionados
    pagos = cliente.pagos.all().order_by('-fecha_vencimiento')[:5]
    
    # Obtener historial de cambios
    historial = cliente.history.all()[:10] if hasattr(cliente, 'history') else []
    
    context = {
        'cliente': cliente,
        'instalaciones': instalaciones,
        'pagos': pagos,
        'historial': historial,
    }
    
    return render(request, 'clientes/cliente_detail.html', context)


@login_required
def cliente_create(request):
    """Crea un nuevo cliente."""
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            try:
                cliente = form.save(commit=False)
                cliente.save(user=request.user)  # Pasar usuario para auditoría
                messages.success(request, f'Cliente "{cliente.nombre_completo}" creado exitosamente.')
                return redirect('clientes:cliente_detail', pk=cliente.pk)
            except Exception as e:
                messages.error(request, f'Error al guardar el cliente: {str(e)}')
        else:
            # Mostrar errores del formulario
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = ClienteForm()
    
    context = {
        'form': form,
        'title': 'Nuevo Cliente',
    }
    
    return render(request, 'clientes/cliente_form.html', context)


@login_required
def cliente_update(request, pk):
    """Actualiza un cliente existente."""
    cliente = get_object_or_404(Cliente.all_objects, pk=pk)  # Usar all_objects para permitir editar eliminados
    
    if cliente.is_deleted:
        messages.warning(request, 'Este cliente está eliminado. Será restaurado al guardar.')
    
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            try:
                cliente = form.save(commit=False)
                # Si estaba eliminado, restaurarlo
                if cliente.is_deleted:
                    cliente.restore(user=request.user)
                else:
                    cliente.save(user=request.user)  # Pasar usuario para auditoría
                messages.success(request, f'Cliente "{cliente.nombre_completo}" actualizado exitosamente.')
                return redirect('clientes:cliente_detail', pk=cliente.pk)
            except Exception as e:
                messages.error(request, f'Error al guardar el cliente: {str(e)}')
        else:
            # Mostrar errores del formulario
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = ClienteForm(instance=cliente)
    
    context = {
        'form': form,
        'cliente': cliente,
        'title': f'Editar Cliente: {cliente.nombre_completo}',
    }
    
    return render(request, 'clientes/cliente_form.html', context)


@login_required
def cliente_delete(request, pk):
    """Elimina un cliente (soft delete)."""
    cliente = get_object_or_404(Cliente.all_objects, pk=pk)
    
    if cliente.is_deleted:
        messages.warning(request, 'Este cliente ya está eliminado.')
        return redirect('clientes:cliente_list')
    
    if request.method == 'POST':
        nombre_completo = cliente.nombre_completo
        cliente.delete(user=request.user)  # Soft delete con usuario
        messages.success(request, f'Cliente "{nombre_completo}" eliminado exitosamente.')
        return redirect('clientes:cliente_list')
    
    context = {
        'cliente': cliente,
    }
    
    return render(request, 'clientes/cliente_confirm_delete.html', context)


@login_required
def cliente_restore(request, pk):
    """Restaura un cliente eliminado."""
    cliente = get_object_or_404(Cliente.all_objects, pk=pk)
    
    if not cliente.is_deleted:
        messages.warning(request, 'Este cliente no está eliminado.')
        return redirect('clientes:cliente_detail', pk=cliente.pk)
    
    if request.method == 'POST':
        nombre_completo = cliente.nombre_completo
        cliente.restore(user=request.user)
        messages.success(request, f'Cliente "{nombre_completo}" restaurado exitosamente.')
        return redirect('clientes:cliente_detail', pk=cliente.pk)
    
    context = {
        'cliente': cliente,
    }
    
    return render(request, 'clientes/cliente_confirm_restore.html', context)


@login_required
@require_http_methods(["POST"])
def cliente_crear_usuario_portal(request, pk):
    """Crea un usuario para el portal de un cliente y muestra las credenciales."""
    cliente = get_object_or_404(Cliente.all_objects, pk=pk)
    
    # Verificar si ya tiene usuario
    if cliente.usuario:
        messages.info(request, f'El cliente "{cliente.nombre_completo}" ya tiene un usuario asignado: {cliente.usuario.username}')
        return redirect('clientes:cliente_detail', pk=cliente.pk)
    
    # Verificar que no esté eliminado
    if cliente.is_deleted:
        messages.error(request, 'No se puede crear usuario para un cliente eliminado.')
        return redirect('clientes:cliente_detail', pk=cliente.pk)
    
    try:
        # Crear usuario (generará contraseña y enviará email automáticamente)
        usuario = cliente.crear_usuario_portal(enviar_email=True)
        
        # Mensaje de éxito
        messages.success(
            request,
            f'✅ Usuario creado exitosamente para "{cliente.nombre_completo}". '
            f'Las credenciales se han enviado por email a {cliente.email}'
        )
        
        # Guardar información en la sesión para mostrar en el template
        request.session['credenciales_cliente'] = {
            'cliente': cliente.nombre_completo,
            'username': usuario.username,
            'email': cliente.email,
            'telefono': cliente.telefono,
            'email_enviado': True,
        }
        # Marcar la sesión como modificada
        request.session.modified = True
        
    except Exception as e:
        messages.error(request, f'Error al crear usuario: {str(e)}')
        # Limpiar credenciales si hay error
        if 'credenciales_cliente' in request.session:
            del request.session['credenciales_cliente']
    
    return redirect('clientes:cliente_detail', pk=cliente.pk)


@login_required
def cliente_verificar_duplicado(request):
    """API endpoint para verificar duplicados en tiempo real."""
    campo = request.GET.get('campo')
    valor = request.GET.get('valor', '').strip()
    cliente_id = request.GET.get('cliente_id')
    
    if not campo:
        return JsonResponse({'error': 'Campo requerido'}, status=400)
    
    if campo not in ['email', 'telefono']:
        return JsonResponse({'error': 'Campo inválido'}, status=400)
    
    # Si el valor está vacío, no validar (pero retornar válido)
    if not valor:
        return JsonResponse({
            'valido': True,
            'mensaje': f'{campo.capitalize()} disponible'
        })
    
    # Validar formato básico para email
    if campo == 'email':
        from django.core.validators import validate_email
        from django.core.exceptions import ValidationError
        try:
            validate_email(valor)
        except ValidationError:
            return JsonResponse({
                'valido': False,
                'mensaje': 'Formato de email inválido'
            })
    
    # Buscar duplicados (solo en clientes no eliminados)
    qs = Cliente.objects.filter(**{campo: valor})
    
    # Excluir el cliente actual si se está editando
    if cliente_id:
        qs = qs.exclude(pk=cliente_id)
    
    existe = qs.exists()
    
    return JsonResponse({
        'valido': not existe,
        'mensaje': f'Ya existe un cliente con este {campo}' if existe else f'{campo.capitalize()} disponible'
    })


@login_required
def cliente_exportar_excel(request):
    """Exporta la lista de clientes a Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    
    # Obtener clientes (aplicar mismos filtros que la lista)
    mostrar_eliminados = request.GET.get('mostrar_eliminados', '') == '1'
    
    if mostrar_eliminados:
        clientes = Cliente.all_objects.all()
    else:
        clientes = Cliente.objects.all()
    
    # Aplicar filtros de búsqueda
    query = request.GET.get('q', '')
    if query:
        clientes = clientes.filter(
            Q(nombre__icontains=query) |
            Q(apellido1__icontains=query) |
            Q(apellido2__icontains=query) |
            Q(telefono__icontains=query) |
            Q(email__icontains=query) |
            Q(ciudad__icontains=query)
        )
    
    estado_filter = request.GET.get('estado', '')
    if estado_filter:
        clientes = clientes.filter(estado_cliente=estado_filter)
    
    orden = request.GET.get('orden', '-created_at')
    if orden in ['nombre', '-nombre']:
        if orden == 'nombre':
            clientes = clientes.order_by('apellido1', 'apellido2', 'nombre')
        else:
            clientes = clientes.order_by('-apellido1', '-apellido2', '-nombre')
    else:
        clientes = clientes.order_by(orden)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Encabezados
    headers = ['Nombre', 'Apellido 1', 'Apellido 2', 'Email', 'Teléfono', 'Dirección', 'Ciudad', 'Estado', 'Estado Cliente', 'Fecha Registro']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos
    for row, cliente in enumerate(clientes, 2):
        ws.cell(row=row, column=1, value=cliente.nombre)
        ws.cell(row=row, column=2, value=cliente.apellido1)
        ws.cell(row=row, column=3, value=cliente.apellido2 or '')
        ws.cell(row=row, column=4, value=cliente.email or '')
        ws.cell(row=row, column=5, value=cliente.telefono)
        ws.cell(row=row, column=6, value=cliente.direccion)
        ws.cell(row=row, column=7, value=cliente.ciudad)
        ws.cell(row=row, column=8, value=cliente.estado)
        ws.cell(row=row, column=9, value=cliente.get_estado_cliente_display())
        ws.cell(row=row, column=10, value=cliente.created_at.strftime('%d/%m/%Y %H:%M') if cliente.created_at else '')
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'clientes_export_{date.today().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def cliente_exportar_pdf(request):
    """Exporta la lista de clientes a PDF."""
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from django.http import HttpResponse
    
    # Obtener clientes (aplicar mismos filtros que la lista)
    mostrar_eliminados = request.GET.get('mostrar_eliminados', '') == '1'
    
    if mostrar_eliminados:
        clientes = Cliente.all_objects.all()
    else:
        clientes = Cliente.objects.all()
    
    # Aplicar filtros
    query = request.GET.get('q', '')
    if query:
        clientes = clientes.filter(
            Q(nombre__icontains=query) |
            Q(apellido1__icontains=query) |
            Q(apellido2__icontains=query) |
            Q(telefono__icontains=query) |
            Q(email__icontains=query) |
            Q(ciudad__icontains=query)
        )
    
    estado_filter = request.GET.get('estado', '')
    if estado_filter:
        clientes = clientes.filter(estado_cliente=estado_filter)
    
    orden = request.GET.get('orden', '-created_at')
    if orden in ['nombre', '-nombre']:
        if orden == 'nombre':
            clientes = clientes.order_by('apellido1', 'apellido2', 'nombre')
        else:
            clientes = clientes.order_by('-apellido1', '-apellido2', '-nombre')
    else:
        clientes = clientes.order_by(orden)
    
    # Crear respuesta PDF
    response = HttpResponse(content_type='application/pdf')
    filename = f'clientes_export_{date.today().strftime("%Y%m%d")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=30,
    )
    
    # Título
    elements.append(Paragraph("Lista de Clientes", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Datos de la tabla
    data = [['Nombre', 'Email', 'Teléfono', 'Ciudad', 'Estado', 'Fecha Registro']]
    
    for cliente in clientes:
        data.append([
            cliente.nombre_completo,
            cliente.email or 'N/A',
            cliente.telefono,
            cliente.ciudad,
            cliente.get_estado_cliente_display(),
            cliente.created_at.strftime('%d/%m/%Y') if cliente.created_at else ''
        ])
    
    # Crear tabla
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    return response


@login_required
def cliente_importar_excel(request):
    """Importa clientes desde un archivo Excel."""
    from openpyxl import load_workbook
    from django.core.exceptions import ValidationError
    
    if request.method == 'POST':
        if 'archivo' not in request.FILES:
            messages.error(request, 'No se seleccionó ningún archivo.')
            return render(request, 'clientes/cliente_importar.html')
        
        archivo = request.FILES['archivo']
        
        try:
            wb = load_workbook(archivo)
            ws = wb.active
            
            creados = 0
            errores = 0
            errores_detalle = []
            
            # Leer filas (saltar encabezado)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not any(row):  # Fila vacía
                    continue
                
                try:
                    # Mapear columnas (ajustar según tu formato)
                    nombre = str(row[0]).strip() if row[0] else ''
                    apellido1 = str(row[1]).strip() if row[1] else ''
                    apellido2 = str(row[2]).strip() if row[2] else None
                    email = str(row[3]).strip() if row[3] else None
                    telefono = str(row[4]).strip() if row[4] else ''
                    direccion = str(row[5]).strip() if row[5] else ''
                    ciudad = str(row[6]).strip() if row[6] else ''
                    estado = str(row[7]).strip() if row[7] else ''
                    
                    # Validaciones básicas
                    if not nombre or not apellido1 or not telefono:
                        errores += 1
                        errores_detalle.append(f'Fila {row_idx}: Faltan datos requeridos')
                        continue
                    
                    # Crear cliente
                    cliente = Cliente(
                        nombre=nombre,
                        apellido1=apellido1,
                        apellido2=apellido2,
                        email=email,
                        telefono=telefono,
                        direccion=direccion,
                        ciudad=ciudad,
                        estado=estado,
                        estado_cliente='activo'
                    )
                    cliente.save(user=request.user)
                    creados += 1
                    
                except Exception as e:
                    errores += 1
                    errores_detalle.append(f'Fila {row_idx}: {str(e)}')
            
            if creados > 0:
                messages.success(request, f'{creados} cliente(s) importado(s) exitosamente.')
            if errores > 0:
                messages.warning(request, f'{errores} error(es) durante la importación.')
                if len(errores_detalle) <= 10:
                    for error in errores_detalle:
                        messages.warning(request, error)
            
            return redirect('clientes:cliente_list')
            
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
    
    return render(request, 'clientes/cliente_importar.html')


@login_required
@require_http_methods(["POST"])
def cliente_bulk_action(request):
    """Realiza acciones masivas sobre clientes seleccionados."""
    action = request.POST.get('action')
    cliente_ids = request.POST.getlist('cliente_ids')
    
    if not action or not cliente_ids:
        messages.error(request, 'Debes seleccionar una acción y al menos un cliente.')
        return redirect('clientes:cliente_list')
    
    clientes = Cliente.all_objects.filter(pk__in=cliente_ids)
    
    if action == 'cambiar_estado':
        nuevo_estado = request.POST.get('nuevo_estado')
        if nuevo_estado:
            actualizados = 0
            for cliente in clientes:
                cliente.estado_cliente = nuevo_estado
                cliente.save(user=request.user)
                actualizados += 1
            messages.success(request, f'{actualizados} cliente(s) actualizado(s) exitosamente.')
    
    elif action == 'eliminar':
        eliminados = 0
        for cliente in clientes:
            if not cliente.is_deleted:
                cliente.delete(user=request.user)
                eliminados += 1
        messages.success(request, f'{eliminados} cliente(s) eliminado(s) exitosamente.')
    
    elif action == 'restaurar':
        restaurados = 0
        for cliente in clientes.filter(is_deleted=True):
            cliente.restore(user=request.user)
            restaurados += 1
        messages.success(request, f'{restaurados} cliente(s) restaurado(s) exitosamente.')
    
    return redirect('clientes:cliente_list')
